"""Tests for shortlist export (Phase U5) — unit + server tool."""

import csv
import os

import pytest

from resume_screener.export import write_csv, write_xlsx
from resume_screener.server import export_shortlist

HERE = os.path.dirname(os.path.abspath(__file__))
SAMPLES = os.path.join(HERE, "sample_resumes")

ROWS = [
    {"rank": 1, "filename": "alice_backend.pdf", "score": 92, "reason": "Strong Python/AWS"},
    {"rank": 2, "filename": "henry_principal.pdf", "score": 88, "reason": "12 yrs, led platform"},
]


@pytest.fixture(scope="session", autouse=True)
def ensure_samples():
    if not os.path.isdir(SAMPLES) or not os.listdir(SAMPLES):
        import generate_samples

        generate_samples.main()


# --- Unit: writers ----------------------------------------------------------

def test_write_csv_roundtrip(tmp_path):
    out = write_csv(ROWS, str(tmp_path / "shortlist.csv"))
    assert os.path.isfile(out)
    with open(out, newline="", encoding="utf-8-sig") as f:
        read = list(csv.reader(f))
    assert read[0] == ["rank", "filename", "score", "reason"]
    assert read[1][1] == "alice_backend.pdf"
    assert read[2][2] == "88"


def test_write_xlsx_roundtrip(tmp_path):
    from openpyxl import load_workbook

    out = write_xlsx(ROWS, str(tmp_path / "shortlist.xlsx"))
    assert os.path.isfile(out)
    wb = load_workbook(out)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == ["rank", "filename", "score", "reason"]
    assert ws.cell(row=2, column=2).value == "alice_backend.pdf"


def test_columns_include_extra_keys(tmp_path):
    rows = [{"filename": "a.pdf", "score": 80, "custom_note": "x"}]
    out = write_csv(rows, str(tmp_path / "x.csv"))
    with open(out, newline="", encoding="utf-8-sig") as f:
        header = next(csv.reader(f))
    assert "custom_note" in header  # extra keys are not dropped


# --- Server tool ------------------------------------------------------------

def test_export_tool_csv(tmp_path):
    out = export_shortlist(ROWS, str(tmp_path / "s.csv"), fmt="csv")
    assert out["ok"] and out["rows"] == 2 and out["format"] == "csv"
    assert os.path.isfile(out["path"])


def test_export_tool_xlsx(tmp_path):
    out = export_shortlist(ROWS, str(tmp_path / "s.xlsx"), fmt="xlsx")
    assert out["ok"] and os.path.isfile(out["path"])


def test_export_bad_format(tmp_path):
    out = export_shortlist(ROWS, str(tmp_path / "s.txt"), fmt="txt")
    assert out["ok"] is False and "format" in out["error"].lower()


def test_export_empty_results(tmp_path):
    assert export_shortlist([], str(tmp_path / "s.csv"))["ok"] is False


def test_export_missing_filename(tmp_path):
    out = export_shortlist([{"score": 90}], str(tmp_path / "s.csv"))
    assert out["ok"] is False and "filename" in out["error"].lower()


def test_export_missing_folder():
    out = export_shortlist(ROWS, os.path.join(HERE, "nope_dir", "s.csv"))
    assert out["ok"] is False and "doesn't exist" in out["error"].lower()


def test_export_enrich_fills_email(tmp_path):
    # alice has an email in her resume; the row omits it -> enrichment fills it.
    rows = [{"filename": "alice_backend.pdf", "score": 90, "reason": "x"}]
    out_path = str(tmp_path / "enriched.csv")
    res = export_shortlist(rows, out_path, fmt="csv", folder=SAMPLES)
    assert res["ok"] and res["enriched"] is True
    with open(res["path"], newline="", encoding="utf-8-sig") as f:
        reader = list(csv.reader(f))
    header = reader[0]
    assert "email" in header
    email_idx = header.index("email")
    assert "alice.johnson@example.com" in reader[1][email_idx]
